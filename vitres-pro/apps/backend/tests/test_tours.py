import json
import unittest
from datetime import date, time, timedelta
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

from fastapi import HTTPException
from openpyxl import load_workbook

from app.routers.tours import (
    _assert_run_access,
    _validate_template_activation,
    _workweek_buckets,
    _xlsx_bytes,
    schedule_is_due,
    service_is_due,
)
from app.schemas.schemas import TourTemplateInput


ROOT = Path(__file__).resolve().parents[1]


def rule(
    *,
    kind="interval",
    anchor=date(2026, 1, 5),
    interval=1,
    months=None,
    cap=None,
):
    return SimpleNamespace(
        kind=kind,
        anchor_date=anchor,
        interval_weeks=interval,
        active_months=months or list(range(1, 13)),
        monthly_cap=cap,
    )


class FrequencyTests(unittest.TestCase):
    def test_supported_week_intervals_stay_anchored(self):
        anchor = date(2026, 1, 5)
        for interval in (1, 2, 4, 6, 8, 12):
            with self.subTest(interval=interval):
                schedule = rule(anchor=anchor, interval=interval)
                self.assertTrue(schedule_is_due(schedule, anchor + timedelta(weeks=interval)))
                not_due = anchor + (timedelta(days=1) if interval == 1 else timedelta(weeks=interval + 1))
                self.assertFalse(schedule_is_due(schedule, not_due))

    def test_missed_service_does_not_move_the_cycle(self):
        schedule = rule(anchor=date(2026, 1, 6), interval=2)
        self.assertTrue(schedule_is_due(schedule, date(2026, 1, 20)))
        # Le resultat terrain du 20 janvier n'entre volontairement jamais dans
        # la fonction : la prochaine date reste ancree au 3 fevrier.
        self.assertFalse(schedule_is_due(schedule, date(2026, 1, 27)))
        self.assertTrue(schedule_is_due(schedule, date(2026, 2, 3)))

    def test_on_demand_is_never_suggested(self):
        self.assertFalse(schedule_is_due(rule(kind="on_demand", anchor=None, interval=None), date(2026, 8, 15)))

    def test_annual_uses_anchor_iso_week(self):
        schedule = rule(kind="annual", anchor=date(2025, 6, 10), interval=None)
        same_week = date.fromisocalendar(2026, date(2025, 6, 10).isocalendar().week, 2)
        self.assertTrue(schedule_is_due(schedule, same_week))
        self.assertFalse(schedule_is_due(schedule, same_week + timedelta(weeks=1)))

    def test_monthly_cap_excludes_fifth_week(self):
        schedule = rule(anchor=date(2026, 3, 2), interval=1, cap=4)
        mondays = [date(2026, 3, 2) + timedelta(weeks=index) for index in range(5)]
        self.assertEqual([schedule_is_due(schedule, day) for day in mondays], [True, True, True, True, False])

    def test_multiple_seasonal_rules(self):
        service = SimpleNamespace(
            active=True,
            needs_review=False,
            schedules=[
                rule(anchor=date(2026, 1, 5), interval=2, months=[4, 5, 6, 7, 8, 9, 10]),
                rule(anchor=date(2026, 1, 5), interval=4, months=[1, 2, 3, 11, 12]),
            ],
        )
        self.assertTrue(service_is_due(service, date(2026, 5, 11)))
        self.assertTrue(service_is_due(service, date(2026, 2, 2)))
        self.assertFalse(service_is_due(service, date(2026, 2, 16)))


class PermissionTests(unittest.TestCase):
    def setUp(self):
        assigned = SimpleNamespace(id="employee-1")
        intervention = SimpleNamespace(employees=[assigned], status="planned")
        self.run = SimpleNamespace(publication_status="published", intervention=intervention)

    def test_assigned_employee_can_execute(self):
        _assert_run_access(self.run, SimpleNamespace(id="employee-1", role="employee"), execute=True)

    def test_unassigned_employee_is_refused(self):
        with self.assertRaises(HTTPException) as context:
            _assert_run_access(self.run, SimpleNamespace(id="employee-2", role="employee"))
        self.assertEqual(context.exception.status_code, 403)

    def test_subcontractor_is_refused(self):
        with self.assertRaises(HTTPException) as context:
            _assert_run_access(self.run, SimpleNamespace(id="sub-1", role="subcontractor"))
        self.assertEqual(context.exception.status_code, 403)

    def test_draft_is_invisible_to_employee(self):
        self.run.publication_status = "draft"
        with self.assertRaises(HTTPException):
            _assert_run_access(self.run, SimpleNamespace(id="employee-1", role="employee"))


class TemplateValidationTests(unittest.TestCase):
    def _payload(self, **changes):
        values = {
            "name": "Tournee test",
            "zone": "hainaut",
            "weekday": 2,
            "default_start_time": time(8),
            "default_end_time": time(16),
        }
        values.update(changes)
        return TourTemplateInput(**values)

    def test_archived_template_cannot_be_active(self):
        with self.assertRaises(HTTPException) as context:
            _validate_template_activation(self._payload(active=True, archived=True, setup_complete=True))
        self.assertEqual(context.exception.status_code, 422)

    def test_default_end_must_follow_start(self):
        with self.assertRaises(HTTPException) as context:
            _validate_template_activation(self._payload(default_start_time=time(16), default_end_time=time(8)))
        self.assertEqual(context.exception.status_code, 422)


class BillingWorkbookTests(unittest.TestCase):
    def test_five_workweek_columns(self):
        self.assertEqual(len(_workweek_buckets(date(2026, 3, 1))), 5)
        self.assertEqual(len(_workweek_buckets(date(2026, 2, 1))), 5)

    def test_workbook_names_numeric_amounts_and_totals(self):
        payload = {
            "monthly": {
                "headers": ["S1", "S2", "S3", "S4", "S5"],
                "rows": [{"export_label": "Commerce A", "amounts": [10.5, 20, 0, 5, 4.5]}],
            },
            "quarterly": {
                "headers": [f"S{index}" for index in range(1, 16)],
                "rows": [{"export_label": "Commerce B", "amounts": [1] * 15}],
            },
        }
        content = _xlsx_bytes(payload)
        workbook = load_workbook(BytesIO(content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Facturation", "Facturation trimestrielle"])
        monthly = workbook["Facturation"]
        self.assertIsInstance(monthly["B2"].value, (int, float))
        self.assertEqual(monthly["B2"].value, 10.5)
        self.assertEqual(monthly["G2"].value, "=SUM(B2:F2)")
        quarterly = workbook["Facturation trimestrielle"]
        self.assertEqual(quarterly["Q2"].value, "=SUM(B2:P2)")


class InitialSeedTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        seed_path = ROOT / "data" / "tours_initial_seed.json"
        if not seed_path.exists():
            raise unittest.SkipTest("La reprise privée Word n'est pas présente dans cet environnement")
        cls.seed = json.loads(seed_path.read_text(encoding="utf-8"))

    def test_all_source_documents_are_present_without_ignored_rows(self):
        self.assertEqual(self.seed["stats"]["templates"], 21)
        self.assertEqual(self.seed["stats"]["hainaut"], 11)
        self.assertEqual(self.seed["stats"]["ardennes"], 10)
        self.assertEqual(self.seed["stats"]["stops"], 517)
        self.assertEqual(self.seed["stats"]["services"], 875)
        self.assertTrue(all(not template["import_report"]["ignored_rows"] for template in self.seed["templates"]))

    def test_mons_is_wednesday_and_two_ardennes_route_two_are_distinct(self):
        mons = next(template for template in self.seed["templates"] if template["name"] == "Tournée Mons")
        self.assertEqual(mons["weekday"], 3)
        wednesday_route_twos = [
            template for template in self.seed["templates"]
            if template["zone"] == "ardennes" and template["weekday"] == 3 and template["name"].startswith("Tournée 2")
        ]
        self.assertEqual(len(wednesday_route_twos), 2)

    def test_import_is_safe_by_default(self):
        self.assertTrue(all(not template["active"] and not template["setup_complete"] for template in self.seed["templates"]))
        ambiguous = [
            service
            for template in self.seed["templates"]
            for section in template["sections"]
            for stop in section["stops"]
            for service in stop["services"]
            if service["needs_review"]
        ]
        self.assertTrue(ambiguous)
        self.assertTrue(all(not service["active"] for service in ambiguous))

    def test_four_structured_billing_modes_are_represented(self):
        modes = {
            service["billing_mode"]
            for template in self.seed["templates"]
            for section in template["sections"]
            for stop in section["stops"]
            for service in stop["services"]
        }
        self.assertEqual(modes, {"monthly_invoice", "quarterly_invoice", "cash_invoiced", "cash_no_invoice"})


if __name__ == "__main__":
    unittest.main()
