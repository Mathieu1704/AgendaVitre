from __future__ import annotations

from collections import defaultdict
from calendar import monthrange
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional, Tuple
from uuid import UUID
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, selectinload

from app.core.deps import get_current_user
from app.core.idempotency import already_processed, record_operation
from app.models.models import (
    Employee,
    Intervention,
    TourBillingBatch,
    TourBillingReview,
    TourRun,
    TourRunCash,
    TourRunService,
    TourRunStop,
    TourSection,
    TourService,
    TourServiceSchedule,
    TourStop,
    TourTemplate,
    get_db,
)
from app.schemas.schemas import (
    TourBillingBatchOut,
    TourBillingExportInput,
    TourBillingReviewInput,
    TourCashInput,
    TourDraftGenerateInput,
    TourDraftScheduleInput,
    TourDraftSelectionInput,
    TourPublishInput,
    TourRunOut,
    TourServiceStatusInput,
    TourStopResolveInput,
    TourTemplateInput,
    TourTemplateOut,
)


router = APIRouter()
BRUSSELS = ZoneInfo("Europe/Brussels")
CASH_MODES = {"cash_invoiced", "cash_no_invoice"}


def _admin(user: Employee) -> None:
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Action reservee aux administrateurs.")


def _template_query(db: Session):
    return db.query(TourTemplate).options(
        selectinload(TourTemplate.sections)
        .selectinload(TourSection.stops)
        .selectinload(TourStop.services)
        .selectinload(TourService.schedules),
    )


def _run_query(db: Session):
    return db.query(TourRun).options(
        selectinload(TourRun.intervention).selectinload(Intervention.employees),
        selectinload(TourRun.stops).selectinload(TourRunStop.services),
        selectinload(TourRun.stops).selectinload(TourRunStop.cash_confirmations),
    )


def _load_run(db: Session, run_id: UUID, for_update: bool = False) -> TourRun:
    query = _run_query(db).filter(TourRun.id == run_id)
    if for_update:
        query = query.with_for_update()
    run = query.first()
    if not run:
        raise HTTPException(status_code=404, detail="Tournee introuvable.")
    return run


def _assert_run_access(run: TourRun, user: Employee, execute: bool = False) -> None:
    if user.role == "admin":
        return
    if user.role != "employee":
        raise HTTPException(status_code=403, detail="Les sous-traitants n'ont pas acces aux tournees.")
    if run.publication_status != "published":
        raise HTTPException(status_code=403, detail="Ce brouillon n'est pas publie.")
    assigned = any(employee.id == user.id for employee in run.intervention.employees)
    if not assigned:
        raise HTTPException(status_code=403, detail="Vous n'etes pas assigne a cette tournee.")
    if execute and run.intervention.status in {"done", "cancelled"}:
        raise HTTPException(status_code=409, detail="La tournee doit etre reouverte avant correction.")


def _local_datetime(day: date, value: time) -> datetime:
    return datetime.combine(day, value, tzinfo=BRUSSELS).astimezone(timezone.utc)


def _month_bounds(day: date) -> Tuple[date, date]:
    start = day.replace(day=1)
    end = day.replace(day=monthrange(day.year, day.month)[1])
    return start, end


def _quarter_bounds(day: date) -> Tuple[date, date]:
    month = ((day.month - 1) // 3) * 3 + 1
    start = date(day.year, month, 1)
    end_month = month + 2
    end = date(day.year, end_month, monthrange(day.year, end_month)[1])
    return start, end


def _workweek_buckets(month_start: date) -> List[date]:
    """Cinq semaines de travail qui couvrent le mois (heure de Bruxelles).

    Une semaine situee avant le mois n'est gardee que si elle contient au moins
    un jour ouvre du mois. Les mois ne contiennent jamais plus de cinq groupes
    lundi-vendredi; un cinquieme groupe vide est ajoute pour les mois courts.
    """
    month_end = _month_bounds(month_start)[1]
    cursor = month_start - timedelta(days=month_start.weekday())
    buckets: List[date] = []
    while cursor <= month_end:
        work_days = [cursor + timedelta(days=i) for i in range(5)]
        if any(month_start <= value <= month_end for value in work_days):
            buckets.append(cursor)
        cursor += timedelta(days=7)
    while len(buckets) < 5:
        buckets.append((buckets[-1] if buckets else month_start) + timedelta(days=7))
    return buckets[:5]


def _bucket_for(day: date, month_start: date) -> date:
    buckets = _workweek_buckets(month_start)
    monday = day - timedelta(days=day.weekday())
    if monday in buckets:
        return monday
    return min(buckets, key=lambda value: abs((value - monday).days))


def schedule_is_due(schedule: TourServiceSchedule, scheduled_date: date) -> bool:
    """Regle pure: une prestation ratee ne decale jamais le cycle d'ancrage."""
    if schedule.kind == "on_demand" or not schedule.anchor_date:
        return False
    months = schedule.active_months or list(range(1, 13))
    if scheduled_date.month not in months or scheduled_date < schedule.anchor_date:
        return False
    if schedule.kind == "annual":
        return scheduled_date.isocalendar().week == schedule.anchor_date.isocalendar().week
    interval = schedule.interval_weeks
    if not interval:
        return False
    days = (scheduled_date - schedule.anchor_date).days
    if days % (interval * 7) != 0:
        return False
    if schedule.monthly_cap:
        month_start, _ = _month_bounds(scheduled_date)
        cursor = month_start
        due_before = 0
        while cursor <= scheduled_date:
            if cursor.weekday() == scheduled_date.weekday():
                elapsed = (cursor - schedule.anchor_date).days
                if cursor >= schedule.anchor_date and elapsed % (interval * 7) == 0:
                    due_before += 1
            cursor += timedelta(days=1)
        if due_before > schedule.monthly_cap:
            return False
    return True


def service_is_due(service: TourService, scheduled_date: date) -> bool:
    if not service.active or service.needs_review:
        return False
    return any(schedule_is_due(rule, scheduled_date) for rule in service.schedules)


def _replace_template_tree(db: Session, template: TourTemplate, payload: TourTemplateInput) -> None:
    db.query(TourStop).filter(TourStop.template_id == template.id).delete(synchronize_session=False)
    db.query(TourSection).filter(TourSection.template_id == template.id).delete(synchronize_session=False)
    db.flush()

    for section_index, section_data in enumerate(payload.sections):
        section = TourSection(
            template_id=template.id,
            label=section_data.label.strip(),
            position=section_data.position if section_data.position is not None else section_index,
        )
        db.add(section)
        db.flush()
        for stop_index, stop_data in enumerate(section_data.stops):
            name = stop_data.name.strip()
            stop = TourStop(
                template_id=template.id,
                section_id=section.id,
                name=name,
                export_label=(stop_data.export_label or name).strip(),
                address=stop_data.address,
                phone=stop_data.phone,
                email=stop_data.email,
                latitude=stop_data.latitude,
                longitude=stop_data.longitude,
                time_window=stop_data.time_window,
                estimated_minutes=stop_data.estimated_minutes,
                instructions=stop_data.instructions,
                position=stop_data.position if stop_data.position is not None else stop_index,
                active=stop_data.active,
                needs_review=stop_data.needs_review,
                source_data=stop_data.source_data,
            )
            db.add(stop)
            db.flush()
            for service_index, service_data in enumerate(stop_data.services):
                service = TourService(
                    stop_id=stop.id,
                    label=service_data.label.strip(),
                    price_ht=service_data.price_ht,
                    billing_mode=service_data.billing_mode,
                    position=service_data.position if service_data.position is not None else service_index,
                    active=service_data.active,
                    needs_review=service_data.needs_review,
                    source_data=service_data.source_data,
                )
                db.add(service)
                db.flush()
                for rule_index, rule_data in enumerate(service_data.schedules):
                    db.add(TourServiceSchedule(
                        service_id=service.id,
                        kind=rule_data.kind,
                        anchor_date=rule_data.anchor_date,
                        interval_weeks=rule_data.interval_weeks,
                        active_months=rule_data.active_months,
                        monthly_cap=rule_data.monthly_cap,
                        position=rule_data.position if rule_data.position is not None else rule_index,
                    ))


def _validate_template_activation(payload: TourTemplateInput) -> None:
    if payload.default_end_time <= payload.default_start_time:
        raise HTTPException(status_code=422, detail="L'heure de fin doit suivre l'heure de debut.")
    if payload.active and payload.archived:
        raise HTTPException(status_code=422, detail="Un modele archive ne peut pas rester actif.")
    if payload.active and not payload.setup_complete:
        raise HTTPException(status_code=422, detail="Terminez la configuration avant d'activer le modele.")
    if not payload.setup_complete:
        return
    services = [service for section in payload.sections for stop in section.stops if stop.active for service in stop.services if service.active]
    if not services:
        raise HTTPException(status_code=422, detail="Le modele doit contenir au moins une prestation active.")
    for section in payload.sections:
        for stop in section.stops:
            if not stop.active:
                continue
            if stop.needs_review:
                raise HTTPException(status_code=422, detail=f"Le commerce {stop.name} doit encore etre valide.")
            for service in stop.services:
                if not service.active:
                    continue
                if service.needs_review:
                    raise HTTPException(status_code=422, detail=f"La prestation {service.label} doit encore etre validee.")
                if not service.schedules:
                    raise HTTPException(status_code=422, detail=f"Ajoutez une frequence a {service.label}.")
                for rule in service.schedules:
                    if rule.kind != "on_demand" and not rule.active_months:
                        raise HTTPException(status_code=422, detail=f"Selectionnez au moins un mois actif pour {service.label}.")
                    if rule.kind != "on_demand" and not rule.anchor_date:
                        raise HTTPException(status_code=422, detail=f"Date d'ancrage manquante pour {service.label}.")
                    if rule.kind == "interval" and not rule.interval_weeks:
                        raise HTTPException(status_code=422, detail=f"Intervalle manquant pour {service.label}.")


def _snapshot_template(db: Session, template: TourTemplate, run: TourRun) -> None:
    for stop in list(run.stops):
        db.delete(stop)
    db.flush()
    ordered_stops = [
        stop
        for section in sorted(template.sections, key=lambda item: item.position)
        for stop in sorted((item for item in section.stops if item.active), key=lambda item: item.position)
    ]
    selected_total = Decimal("0")
    for run_position, stop in enumerate(ordered_stops):
        run_stop = TourRunStop(
            run_id=run.id,
            source_stop_id=stop.id,
            section_label=stop.section.label if stop.section else None,
            name=stop.name,
            export_label=stop.export_label or stop.name,
            address=stop.address,
            phone=stop.phone,
            email=stop.email,
            latitude=stop.latitude,
            longitude=stop.longitude,
            time_window=stop.time_window,
            estimated_minutes=stop.estimated_minutes,
            instructions=stop.instructions,
            position=run_position,
            selected=False,
        )
        db.add(run_stop)
        db.flush()
        any_selected = False
        for service in sorted((item for item in stop.services if item.active), key=lambda item: item.position):
            due = service_is_due(service, run.scheduled_date)
            db.add(TourRunService(
                run_stop_id=run_stop.id,
                source_service_id=service.id,
                label=service.label,
                price_ht=service.price_ht,
                billing_mode=service.billing_mode,
                position=service.position,
                suggested=due,
                selected=due,
            ))
            any_selected = any_selected or due
            if due:
                selected_total += Decimal(service.price_ht)
        run_stop.selected = any_selected
    run.template_version = template.version
    run.intervention.price_estimated = selected_total


def _ensure_one_draft(db: Session, template: TourTemplate, scheduled_date: date) -> Tuple[TourRun, bool]:
    existing = db.query(TourRun).filter(
        TourRun.template_id == template.id,
        TourRun.scheduled_date == scheduled_date,
    ).first()
    if existing:
        return existing, False
    intervention = Intervention(
        type="tournee",
        title=template.name,
        description="Tournee recurrente preparee depuis un modele.",
        start_time=_local_datetime(scheduled_date, template.default_start_time),
        end_time=_local_datetime(scheduled_date, template.default_end_time),
        status="planned",
        zone=template.zone,
        time_tbd=False,
        payment_mode="invoice",
    )
    db.add(intervention)
    db.flush()
    run = TourRun(
        template_id=template.id,
        intervention_id=intervention.id,
        scheduled_date=scheduled_date,
        template_version=template.version,
        publication_status="draft",
    )
    db.add(run)
    db.flush()
    _snapshot_template(db, template, run)
    return run, True


def ensure_drafts(db: Session, start_date: date, weeks: int = 8) -> int:
    weeks = max(1, min(weeks, 52))
    end_date = start_date + timedelta(weeks=weeks)
    templates = _template_query(db).filter(
        TourTemplate.active.is_(True),
        TourTemplate.archived.is_(False),
        TourTemplate.setup_complete.is_(True),
    ).with_for_update().all()
    created = 0
    for template in templates:
        cursor = start_date
        cursor += timedelta(days=(template.weekday - cursor.isoweekday()) % 7)
        while cursor < end_date:
            _, was_created = _ensure_one_draft(db, template, cursor)
            created += int(was_created)
            cursor += timedelta(days=7)
    return created


def _refresh_future_drafts(db: Session, template: TourTemplate) -> None:
    today = datetime.now(BRUSSELS).date()
    drafts = db.query(TourRun).filter(
        TourRun.template_id == template.id,
        TourRun.publication_status == "draft",
        TourRun.scheduled_date >= today,
    ).all()
    for run in drafts:
        # Si l'admin change le jour fixe, l'ancien brouillon ne doit pas
        # subsister en plus de celui qui sera regenere au nouveau jour.
        if run.scheduled_date.isoweekday() != template.weekday:
            db.delete(run.intervention)
            continue
        intervention = run.intervention
        intervention.title = template.name
        intervention.zone = template.zone
        intervention.start_time = _local_datetime(run.scheduled_date, template.default_start_time)
        intervention.end_time = _local_datetime(run.scheduled_date, template.default_end_time)
        _snapshot_template(db, template, run)


def _delete_future_drafts(db: Session, template: TourTemplate) -> None:
    today = datetime.now(BRUSSELS).date()
    drafts = db.query(TourRun).filter(
        TourRun.template_id == template.id,
        TourRun.publication_status == "draft",
        TourRun.scheduled_date >= today,
    ).all()
    for run in drafts:
        db.delete(run.intervention)


@router.get("/templates", response_model=List[TourTemplateOut])
def list_templates(
    include_archived: bool = False,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    query = _template_query(db)
    if not include_archived:
        query = query.filter(TourTemplate.archived.is_(False))
    return query.order_by(TourTemplate.zone, TourTemplate.weekday, TourTemplate.name).all()


@router.get("/templates/{template_id}", response_model=TourTemplateOut)
def get_template(
    template_id: UUID,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    template = _template_query(db).filter(TourTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Modele introuvable.")
    return template


@router.post("/templates", response_model=TourTemplateOut)
def create_template(
    payload: TourTemplateInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    if payload.setup_complete and not payload.sections:
        raise HTTPException(status_code=422, detail="Un modele actif doit contenir au moins une section.")
    _validate_template_activation(payload)
    template = TourTemplate(
        name=payload.name.strip(),
        zone=payload.zone,
        weekday=payload.weekday,
        default_start_time=payload.default_start_time,
        default_end_time=payload.default_end_time,
        active=payload.active,
        archived=payload.archived,
        setup_complete=payload.setup_complete,
        source_document=payload.source_document,
    )
    db.add(template)
    db.flush()
    _replace_template_tree(db, template, payload)
    if template.active and template.setup_complete and not template.archived:
        ensure_drafts(db, datetime.now(BRUSSELS).date(), 8)
    db.commit()
    return get_template(template.id, db, current_user)


@router.put("/templates/{template_id}", response_model=TourTemplateOut)
def update_template(
    template_id: UUID,
    payload: TourTemplateInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    template = _template_query(db).filter(TourTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Modele introuvable.")
    _validate_template_activation(payload)
    template.name = payload.name.strip()
    template.zone = payload.zone
    template.weekday = payload.weekday
    template.default_start_time = payload.default_start_time
    template.default_end_time = payload.default_end_time
    template.active = payload.active
    template.archived = payload.archived
    template.setup_complete = payload.setup_complete
    template.source_document = payload.source_document
    template.version += 1
    _replace_template_tree(db, template, payload)
    db.flush()
    db.expire(template, ["sections", "stops"])
    template = _template_query(db).filter(TourTemplate.id == template_id).first()
    _refresh_future_drafts(db, template)
    if template.active and template.setup_complete and not template.archived:
        ensure_drafts(db, datetime.now(BRUSSELS).date(), 8)
    else:
        _delete_future_drafts(db, template)
    db.commit()
    return get_template(template.id, db, current_user)


@router.post("/templates/{template_id}/archive", response_model=TourTemplateOut)
def archive_template(
    template_id: UUID,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    template = db.query(TourTemplate).filter(TourTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Modele introuvable.")
    template.archived = True
    template.active = False
    _delete_future_drafts(db, template)
    db.commit()
    return get_template(template.id, db, current_user)


@router.post("/drafts/generate")
def generate_drafts(
    payload: TourDraftGenerateInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    start = payload.start_date or datetime.now(BRUSSELS).date()
    created = ensure_drafts(db, start, payload.weeks)
    db.commit()
    return {"created": created, "weeks": max(1, min(payload.weeks, 52))}


@router.get("/drafts", response_model=List[TourRunOut])
def list_drafts(
    start: Optional[date] = None,
    weeks: int = Query(8, ge=1, le=52),
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    start_date = start or datetime.now(BRUSSELS).date()
    ensure_drafts(db, start_date, weeks)
    db.commit()
    end = start_date + timedelta(weeks=weeks)
    return _run_query(db).filter(
        TourRun.publication_status == "draft",
        TourRun.scheduled_date >= start_date,
        TourRun.scheduled_date < end,
    ).order_by(TourRun.scheduled_date).all()


@router.patch("/runs/{run_id}/services/{service_id}/selection", response_model=TourRunOut)
def select_draft_service(
    run_id: UUID,
    service_id: UUID,
    payload: TourDraftSelectionInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    run = _load_run(db, run_id, for_update=True)
    if run.publication_status != "draft":
        raise HTTPException(status_code=409, detail="Une tournee publiee est figee.")
    service = next((item for stop in run.stops for item in stop.services if item.id == service_id), None)
    if not service:
        raise HTTPException(status_code=404, detail="Prestation introuvable.")
    service.selected = payload.selected
    service.status = "pending"
    service.exception_reason = None
    stop = service.run_stop
    stop.selected = any(item.selected for item in stop.services)
    stop.status = "pending"
    run.intervention.price_estimated = sum(
        (Decimal(item.price_ht) for run_stop in run.stops for item in run_stop.services if item.selected),
        Decimal("0"),
    )
    db.commit()
    return _load_run(db, run_id)


@router.patch("/runs/{run_id}/schedule", response_model=TourRunOut)
def update_draft_schedule(
    run_id: UUID,
    payload: TourDraftScheduleInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    run = _load_run(db, run_id, for_update=True)
    if run.publication_status != "draft":
        raise HTTPException(status_code=409, detail="Une tournee publiee est figee.")
    if payload.end_time <= payload.start_time:
        raise HTTPException(status_code=422, detail="L'heure de fin doit suivre l'heure de debut.")
    if payload.start_time.astimezone(BRUSSELS).date() != run.scheduled_date:
        raise HTTPException(status_code=422, detail="Le jour fixe de cette occurrence ne peut pas etre deplace ici.")
    run.intervention.start_time = payload.start_time
    run.intervention.end_time = payload.end_time
    run.scheduled_date = payload.start_time.astimezone(BRUSSELS).date()
    db.commit()
    return _load_run(db, run_id)


@router.post("/runs/{run_id}/publish", response_model=TourRunOut)
def publish_run(
    run_id: UUID,
    payload: TourPublishInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    run = _load_run(db, run_id, for_update=True)
    if run.publication_status == "published":
        return run
    selected = [service for stop in run.stops for service in stop.services if service.selected]
    if not selected:
        raise HTTPException(status_code=422, detail="Selectionnez au moins une prestation avant publication.")
    if not payload.employee_ids:
        raise HTTPException(status_code=422, detail="Assignez au moins un employe avant publication.")
    employees = db.query(Employee).filter(Employee.id.in_(payload.employee_ids)).all()
    if len(employees) != len(set(payload.employee_ids)):
        raise HTTPException(status_code=422, detail="Un employe assigne est introuvable.")
    invalid = [employee.full_name or employee.email for employee in employees if employee.role != "employee" or employee.zone != run.intervention.zone]
    if invalid:
        raise HTTPException(status_code=422, detail=f"Employes incompatibles avec la tournee: {', '.join(invalid)}")
    run.intervention.employees = employees
    run.publication_status = "published"
    run.published_at = datetime.now(timezone.utc)
    db.commit()
    return _load_run(db, run_id)


@router.get("/assigned", response_model=List[TourRunOut])
def assigned_runs(
    start: Optional[date] = None,
    end: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    if current_user.role == "subcontractor":
        raise HTTPException(status_code=403, detail="Les sous-traitants n'ont pas acces aux tournees.")
    query = _run_query(db).filter(TourRun.publication_status == "published")
    if current_user.role != "admin":
        query = query.filter(TourRun.intervention.has(Intervention.employees.any(id=current_user.id)))
    if start:
        query = query.filter(TourRun.scheduled_date >= start)
    if end:
        query = query.filter(TourRun.scheduled_date <= end)
    return query.order_by(TourRun.scheduled_date).all()


def _touch_in_progress(run: TourRun) -> None:
    if run.intervention.status == "planned":
        run.intervention.status = "in_progress"


def _sync_stop_cash(db: Session, stop: TourRunStop) -> None:
    expected: Dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for service in stop.services:
        if service.selected and service.status == "done" and service.billing_mode in CASH_MODES:
            expected[service.billing_mode] += Decimal(service.price_ht)
    existing = {cash.billing_mode: cash for cash in stop.cash_confirmations}
    for mode in CASH_MODES:
        amount = expected.get(mode, Decimal("0"))
        cash = existing.get(mode)
        if amount <= 0:
            if cash:
                db.delete(cash)
            continue
        if not cash:
            stop.cash_confirmations.append(TourRunCash(billing_mode=mode, expected_amount=amount))
        elif Decimal(cash.expected_amount) != amount:
            cash.expected_amount = amount
            cash.received_amount = None
            cash.confirmed_at = None
            cash.confirmed_by = None


def _derive_stop_status(stop: TourRunStop) -> None:
    services = [service for service in stop.services if service.selected]
    if not services or any(service.status == "pending" for service in services):
        stop.status = "pending"
        stop.completed_at = None
    elif all(service.status == "done" for service in services):
        stop.status = "done"
        stop.exception_reason = None
        stop.completed_at = datetime.now(timezone.utc)
    else:
        stop.status = "partial"
        stop.completed_at = datetime.now(timezone.utc)


@router.patch("/runs/{run_id}/services/{service_id}", response_model=TourRunOut)
def update_run_service(
    run_id: UUID,
    service_id: UUID,
    payload: TourServiceStatusInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    run = _load_run(db, run_id, for_update=True)
    _assert_run_access(run, current_user, execute=True)
    service = next((item for stop in run.stops for item in stop.services if item.id == service_id), None)
    if not service or not service.selected:
        raise HTTPException(status_code=404, detail="Prestation selectionnee introuvable.")
    if payload.status == "not_done" and not (payload.reason or "").strip():
        raise HTTPException(status_code=422, detail="Une justification est obligatoire.")
    service.status = payload.status
    service.exception_reason = payload.reason.strip() if payload.reason else None
    service.performed_at = datetime.now(timezone.utc) if payload.status == "done" else None
    _derive_stop_status(service.run_stop)
    _sync_stop_cash(db, service.run_stop)
    _touch_in_progress(run)
    db.commit()
    return _load_run(db, run_id)


@router.post("/runs/{run_id}/stops/{stop_id}/resolve", response_model=TourRunOut)
def resolve_run_stop(
    run_id: UUID,
    stop_id: UUID,
    payload: TourStopResolveInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    run = _load_run(db, run_id, for_update=True)
    _assert_run_access(run, current_user, execute=True)
    seen = already_processed(db, payload.client_operation_id)
    if seen is not None:
        return run
    stop = next((item for item in run.stops if item.id == stop_id and item.selected), None)
    if not stop:
        raise HTTPException(status_code=404, detail="Commerce selectionne introuvable.")
    services = [service for service in stop.services if service.selected]
    not_done_ids = set(payload.not_done_service_ids)
    now = datetime.now(timezone.utc)
    if payload.status == "not_visited":
        reason = (payload.reason or "").strip()
        if not reason:
            raise HTTPException(status_code=422, detail="Une justification est obligatoire pour un commerce non visite.")
        for service in services:
            service.status = "not_done"
            service.exception_reason = reason
            service.performed_at = None
        stop.status = "not_visited"
        stop.exception_reason = reason
    elif payload.status == "partial":
        if not not_done_ids or len(not_done_ids) >= len(services):
            raise HTTPException(status_code=422, detail="Une tournee partielle doit contenir au moins une prestation faite et une non faite.")
        unknown = not_done_ids - {service.id for service in services}
        if unknown:
            raise HTTPException(status_code=422, detail="Une prestation ne fait pas partie de ce commerce.")
        for service in services:
            if service.id in not_done_ids:
                reason = (payload.service_reasons.get(str(service.id)) or payload.reason or "").strip()
                if not reason:
                    raise HTTPException(status_code=422, detail=f"Justification manquante pour {service.label}.")
                service.status = "not_done"
                service.exception_reason = reason
                service.performed_at = None
            else:
                service.status = "done"
                service.exception_reason = None
                service.performed_at = now
        stop.status = "partial"
        stop.exception_reason = (payload.reason or "").strip() or None
    else:
        for service in services:
            service.status = "done"
            service.exception_reason = None
            service.performed_at = now
        stop.status = "done"
        stop.exception_reason = None
    stop.completed_at = now
    _sync_stop_cash(db, stop)
    db.flush()
    for cash in stop.cash_confirmations:
        if cash.billing_mode in payload.cash_received:
            amount = payload.cash_received[cash.billing_mode]
            if amount < 0:
                raise HTTPException(status_code=422, detail="Le montant encaisse ne peut pas etre negatif.")
            cash.received_amount = amount
            cash.confirmed_at = now
            cash.confirmed_by = current_user.id
    _touch_in_progress(run)
    record_operation(db, payload.client_operation_id, current_user.id, f"POST /api/tours/runs/{run_id}/stops/{stop_id}/resolve", stop.id)
    db.commit()
    return _load_run(db, run_id)


@router.patch("/runs/{run_id}/stops/{stop_id}/cash/{billing_mode}", response_model=TourRunOut)
def confirm_run_cash(
    run_id: UUID,
    stop_id: UUID,
    billing_mode: str,
    payload: TourCashInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    if billing_mode not in CASH_MODES:
        raise HTTPException(status_code=422, detail="Mode d'encaissement invalide.")
    run = _load_run(db, run_id, for_update=True)
    _assert_run_access(run, current_user, execute=True)
    stop = next((item for item in run.stops if item.id == stop_id), None)
    cash = next((item for item in (stop.cash_confirmations if stop else []) if item.billing_mode == billing_mode), None)
    if not cash:
        raise HTTPException(status_code=404, detail="Encaissement attendu introuvable.")
    if payload.received_amount < 0:
        raise HTTPException(status_code=422, detail="Le montant encaisse ne peut pas etre negatif.")
    cash.received_amount = payload.received_amount
    cash.confirmed_at = datetime.now(timezone.utc)
    cash.confirmed_by = current_user.id
    _touch_in_progress(run)
    db.commit()
    return _load_run(db, run_id)


@router.post("/runs/{run_id}/close", response_model=TourRunOut)
def close_run(
    run_id: UUID,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    run = _load_run(db, run_id, for_update=True)
    # Une reponse perdue apres le commit doit pouvoir etre rejouee par
    # l'outbox sans creer un faux echec permanent.
    _assert_run_access(run, current_user)
    if run.intervention.status == "done":
        return run
    if run.intervention.status == "cancelled":
        raise HTTPException(status_code=409, detail="La tournee annulee doit etre reouverte avant cloture.")
    pending = [stop.name for stop in run.stops if stop.selected and stop.status == "pending"]
    if pending:
        raise HTTPException(status_code=409, detail=f"Commerces encore en attente: {', '.join(pending)}")
    missing_cash = [
        f"{stop.name} ({cash.billing_mode})"
        for stop in run.stops if stop.selected
        for cash in stop.cash_confirmations if cash.confirmed_at is None or cash.received_amount is None
    ]
    if missing_cash:
        raise HTTPException(status_code=409, detail=f"Encaissements a confirmer: {', '.join(missing_cash)}")
    run.intervention.status = "done"
    run.completed_at = datetime.now(timezone.utc)
    db.commit()
    return _load_run(db, run_id)


@router.post("/runs/{run_id}/reopen", response_model=TourRunOut)
def reopen_run(
    run_id: UUID,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    run = _load_run(db, run_id, for_update=True)
    if run.publication_status != "published":
        raise HTTPException(status_code=409, detail="Un brouillon n'a pas besoin d'etre reouvert.")
    resolved = any(stop.selected and stop.status != "pending" for stop in run.stops)
    run.intervention.status = "in_progress" if resolved else "planned"
    run.completed_at = None
    db.commit()
    return _load_run(db, run_id)


@router.post("/runs/{run_id}/cancel", response_model=TourRunOut)
def cancel_run(
    run_id: UUID,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    run = _load_run(db, run_id, for_update=True)
    if run.publication_status != "published":
        raise HTTPException(status_code=409, detail="Supprimez simplement le brouillon non publie.")
    run.intervention.status = "cancelled"
    run.completed_at = None
    db.commit()
    return _load_run(db, run_id)


def _exported_source_ids(db: Session, zone: str) -> Tuple[set[str], set[str]]:
    batches = db.query(TourBillingBatch).filter(TourBillingBatch.zone == zone).all()
    services = {value for batch in batches for value in (batch.source_service_ids or [])}
    cash = {value for batch in batches for value in (batch.source_cash_ids or [])}
    return services, cash


def _add_cell(cells, cadence, label, bucket_start, amount, service_id=None, cash_id=None):
    key = (cadence, label, bucket_start)
    cell = cells.setdefault(key, {
        "cadence": cadence,
        "export_label": label,
        "bucket_start": bucket_start,
        "source_amount": Decimal("0"),
        "source_service_ids": [],
        "source_cash_ids": [],
    })
    cell["source_amount"] += Decimal(amount)
    if service_id:
        cell["source_service_ids"].append(str(service_id))
    if cash_id:
        cell["source_cash_ids"].append(str(cash_id))


def _billing_cells(db: Session, zone: str, period_start: date) -> List[dict]:
    period_start = period_start.replace(day=1)
    quarter_start, quarter_end = _quarter_bounds(period_start)
    exported_services, exported_cash = _exported_source_ids(db, zone)
    runs = _run_query(db).join(Intervention, TourRun.intervention_id == Intervention.id).filter(
        TourRun.publication_status == "published",
        TourRun.scheduled_date >= quarter_start,
        TourRun.scheduled_date <= quarter_end,
        Intervention.zone == zone,
        Intervention.status == "done",
    ).all()
    cells: Dict[tuple, dict] = {}
    month_start, month_end = _month_bounds(period_start)
    for run in runs:
        for stop in run.stops:
            if not stop.selected:
                continue
            for service in stop.services:
                service_key = str(service.id)
                if not service.selected or service.status != "done" or service_key in exported_services:
                    continue
                if service.billing_mode == "monthly_invoice" and month_start <= run.scheduled_date <= month_end:
                    _add_cell(cells, "monthly", stop.export_label, _bucket_for(run.scheduled_date, month_start), service.price_ht, service_id=service.id)
                elif service.billing_mode == "quarterly_invoice":
                    service_month = run.scheduled_date.replace(day=1)
                    _add_cell(cells, "quarterly", stop.export_label, _bucket_for(run.scheduled_date, service_month), service.price_ht, service_id=service.id)
            if month_start <= run.scheduled_date <= month_end:
                for cash in stop.cash_confirmations:
                    cash_key = str(cash.id)
                    if cash.billing_mode != "cash_invoiced" or cash.confirmed_at is None or cash_key in exported_cash:
                        continue
                    _add_cell(cells, "monthly", stop.export_label, _bucket_for(run.scheduled_date, month_start), cash.received_amount or 0, cash_id=cash.id)
    reviews = db.query(TourBillingReview).filter(
        TourBillingReview.zone == zone,
        TourBillingReview.period_start == period_start,
    ).all()
    review_map = {(item.cadence, item.export_label, item.bucket_start): item for item in reviews}
    result = []
    for key, cell in cells.items():
        review = review_map.get(key)
        amount = review.override_amount if review and review.override_amount is not None else cell["source_amount"]
        result.append({
            **cell,
            "source_amount": float(cell["source_amount"]),
            "amount": float(amount),
            "selected": review.selected if review else True,
            "overridden": bool(review and review.override_amount is not None),
        })
    return sorted(result, key=lambda item: (item["cadence"], item["export_label"].casefold(), item["bucket_start"]))


def _billing_payload(db: Session, zone: str, period_start: date) -> dict:
    period_start = period_start.replace(day=1)
    quarter_start, _ = _quarter_bounds(period_start)
    cells = _billing_cells(db, zone, period_start)
    monthly_buckets = _workweek_buckets(period_start)
    quarterly_buckets = [bucket for offset in range(3) for bucket in _workweek_buckets(date(quarter_start.year, quarter_start.month + offset, 1))]

    def matrix(cadence: str, buckets: List[date]) -> dict:
        selected = [cell for cell in cells if cell["cadence"] == cadence and cell["selected"]]
        labels = sorted({cell["export_label"] for cell in selected}, key=str.casefold)
        values = {(cell["export_label"], cell["bucket_start"]): cell["amount"] for cell in selected}
        rows = [{
            "export_label": label,
            "amounts": [float(values.get((label, bucket), 0)) for bucket in buckets],
        } for label in labels]
        return {
            "buckets": [bucket.isoformat() for bucket in buckets],
            "headers": [f"S{index % 5 + 1} · {bucket.strftime('%d/%m')} - {(bucket + timedelta(days=4)).strftime('%d/%m')}" for index, bucket in enumerate(buckets)],
            "rows": rows,
        }

    json_cells = [{**cell, "bucket_start": cell["bucket_start"].isoformat()} for cell in cells]
    return {
        "zone": zone,
        "period_start": period_start.isoformat(),
        "monthly": matrix("monthly", monthly_buckets),
        "quarterly": matrix("quarterly", quarterly_buckets),
        "cells": json_cells,
    }


@router.get("/billing/matrix")
def billing_matrix(
    zone: str,
    period_start: date,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    if zone not in {"hainaut", "ardennes"}:
        raise HTTPException(status_code=422, detail="Zone invalide.")
    return _billing_payload(db, zone, period_start)


@router.put("/billing/review")
def save_billing_review(
    zone: str,
    period_start: date,
    payload: TourBillingReviewInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    normalized = period_start.replace(day=1)
    review = db.query(TourBillingReview).filter(
        TourBillingReview.zone == zone,
        TourBillingReview.period_start == normalized,
        TourBillingReview.cadence == payload.cadence,
        TourBillingReview.export_label == payload.export_label,
        TourBillingReview.bucket_start == payload.bucket_start,
    ).first()
    if not review:
        review = TourBillingReview(
            zone=zone,
            period_start=normalized,
            cadence=payload.cadence,
            export_label=payload.export_label,
            bucket_start=payload.bucket_start,
        )
        db.add(review)
    review.selected = payload.selected
    review.override_amount = payload.override_amount
    review.updated_by = current_user.id
    db.commit()
    return {"ok": True}


def _xlsx_bytes(payload: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for sheet_name, key in (("Facturation", "monthly"), ("Facturation trimestrielle", "quarterly")):
        sheet = workbook.create_sheet(sheet_name)
        matrix = payload[key]
        headers = ["Commerce", *matrix["headers"], "Total"]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A5F")
        for row_index, row in enumerate(matrix["rows"], start=2):
            sheet.append([row["export_label"], *row["amounts"], None])
            first = 2
            last = 1 + len(row["amounts"])
            sheet.cell(row=row_index, column=last + 1, value=f"=SUM({sheet.cell(row=row_index, column=first).coordinate}:{sheet.cell(row=row_index, column=last).coordinate})")
            for column in range(first, last + 2):
                sheet.cell(row=row_index, column=column).number_format = '#,##0.00 [$€-fr-BE]'
        sheet.freeze_panes = "B2"
        sheet.column_dimensions["A"].width = 38
        for column in range(2, len(headers) + 1):
            sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = 17
        sheet.auto_filter.ref = sheet.dimensions
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


@router.post("/billing/exports", response_model=TourBillingBatchOut)
def create_billing_export(
    payload: TourBillingExportInput,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    period_start = payload.period_start.replace(day=1)
    export_payload = _billing_payload(db, payload.zone, period_start)
    cells = [cell for cell in export_payload["cells"] if cell["selected"]]
    if not cells:
        raise HTTPException(status_code=422, detail="Aucune cellule selectionnee a exporter.")
    filename = f"facturation_tournees_{payload.zone}_{period_start.strftime('%Y_%m')}.xlsx"
    batch = TourBillingBatch(
        zone=payload.zone,
        period_start=period_start,
        filename=filename,
        payload=export_payload,
        source_service_ids=sorted({value for cell in cells for value in cell["source_service_ids"]}),
        source_cash_ids=sorted({value for cell in cells for value in cell["source_cash_ids"]}),
        created_by=current_user.id,
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return batch


@router.get("/billing/exports", response_model=List[TourBillingBatchOut])
def list_billing_exports(
    zone: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    query = db.query(TourBillingBatch)
    if zone:
        query = query.filter(TourBillingBatch.zone == zone)
    return query.order_by(TourBillingBatch.created_at.desc()).all()


@router.get("/billing/exports/{batch_id}/download")
def download_billing_export(
    batch_id: UUID,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    _admin(current_user)
    batch = db.query(TourBillingBatch).filter(TourBillingBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Lot d'export introuvable.")
    content = _xlsx_bytes(batch.payload)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{batch.filename}"'},
    )


@router.get("/runs/{run_id}", response_model=TourRunOut)
def get_run(
    run_id: UUID,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    run = _load_run(db, run_id)
    _assert_run_access(run, current_user)
    return run


def generate_drafts_job() -> None:
    """Etend l'horizon a huit semaines, sans faire echouer le demarrage API.

    Le verrou transactionnel evite que plusieurs workers generent la meme
    occurrence simultanement. L'unicite template/date reste la seconde garde.
    """
    from sqlalchemy import text
    from app.models.models import SessionLocal

    db = SessionLocal()
    try:
        locked = db.execute(text("SELECT pg_try_advisory_xact_lock(837264021)")).scalar()
        if not locked:
            db.rollback()
            return
        ensure_drafts(db, datetime.now(BRUSSELS).date(), 8)
        db.commit()
    except Exception as error:
        db.rollback()
        # La premiere mise en route peut preceder l'application de la migration.
        # Le prochain passage reprendra automatiquement, sans bloquer l'API.
        print(f"[tour-drafts] generation differee: {error}")
    finally:
        db.close()
